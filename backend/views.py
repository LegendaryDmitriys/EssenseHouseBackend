import django_filters
from django.http import JsonResponse, Http404
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from rest_framework import generics

from rest_framework.generics import ListCreateAPIView
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.core.cache import cache

import openpyxl
from django.http import HttpResponse


from .models import House, ConstructionTechnology, HouseCategory, FinishingOption, Document, Review, Order, \
    UserQuestionHouse, PurchasedHouse, FilterOption, UserQuestion, Image, Blog, BlogCategory, ReviewFile

from .serializer import HouseSerializer, ConstructionTechnologySerializer, HouseCategorySerializer, \
    FinishingOptionSerializer, DocumentSerializer, ReviewSerializer, OrderSerializer, \
    PurchasedHouseSerializer, FilterOptionsSerializer, UserQuestionHouseSerializer, UserQuestionSerializer, \
    BlogSerializer, BlogCategorySerializer
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404



def filter_houses(filters, category=None):
    houses = House.objects.filter(category=category) if category else House.objects.all()

    if 'price_min' in filters and filters['price_min']:
        min_price = int(filters['price_min'])
        houses = houses.filter(price__gte=min_price)

    if 'price_max' in filters and filters['price_max']:
        max_price = int(filters['price_max'])
        houses = houses.filter(price__lte=max_price)

    if 'bestSeller' in filters and filters.getlist('bestSeller'):
        houses = houses.filter(best_seller__in=filters.getlist('bestSeller'))

    if 'area_min' in filters and filters['area_min']:
        min_area = int(filters['area_min'])
        houses = houses.filter(area__gte=min_area)

    if 'area_max' in filters and filters['area_max']:
        max_area = int(filters['area_max'])
        houses = houses.filter(area__lte=max_area)

    if 'floors' in filters and filters.getlist('floors'):
        houses = houses.filter(floors__in=filters.getlist('floors'))

    if 'rooms' in filters and filters.getlist('rooms'):
        houses = houses.filter(rooms__in=filters.getlist('rooms'))

    if 'living_area_min' in filters and filters['living_area_min']:
        min_living_area = int(filters['living_area_min'])
        houses = houses.filter(living_area__gte=min_living_area)

    if 'living_area_max' in filters and filters['living_area_max']:
        max_living_area = int(filters['living_area_max'])
        houses = houses.filter(living_area__lte=max_living_area)

    if 'bedrooms' in filters and filters.getlist('bedrooms'):
        houses = houses.filter(bedrooms__in=filters.getlist('bedrooms'))

    if 'garage' in filters:
        garage_value = True if filters['garage'] == 'Да' else False
        houses = houses.filter(garage=garage_value)

    if 'purpose' in filters and filters.getlist('purpose'):
        houses = houses.filter(purpose__in=filters.getlist('purpose'))

    if 'constructionTechnology' in filters and filters.getlist('constructionTechnology'):
        houses = houses.filter(construction_technology__in=filters.getlist('constructionTechnology'))

    return houses


class DynamicHouseFilter(django_filters.FilterSet):
    class Meta:
        model = House
        fields = {}

class Pagination(PageNumberPagination):
    page_size = 6
    page_size_query_param = 'page_size'
    max_page_size = 100

class HousePagination(PageNumberPagination):
    page_size = 6
    page_size_query_param = 'page_size'
    max_page_size = 100

class HouseListView(APIView):
    serializer_class = HouseSerializer
    parser_classes = [MultiPartParser, FormParser]
    pagination_class = HousePagination

    def get(self, request, id=None):
        if id is not None:
            return self.get_house_by_id(id)

        category_slug = request.query_params.get('category')
        filters = request.query_params
        sort_by = request.query_params.get('sort', 'priceAsc')
        limit = request.query_params.get('limit')
        title = request.query_params.get('title')

        houses = self.filter_houses(filters, category_slug, sort_by, title)

        if limit:
            try:
                limit = int(limit)
                if limit > 0:
                    houses = houses[:limit]
                else:
                    return Response({'detail': 'Параметр limit должен быть положительным числом.'},
                                    status=status.HTTP_400_BAD_REQUEST)
            except ValueError:
                return Response({'detail': 'Параметр limit должен быть целым числом.'},
                                status=status.HTTP_400_BAD_REQUEST)

        if not limit:
            paginator = self.pagination_class()
            paginated_houses = paginator.paginate_queryset(houses, request)
            serializer = self.serializer_class(paginated_houses, many=True)
            return paginator.get_paginated_response(serializer.data)

        serializer = self.serializer_class(houses, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = self.serializer_class(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


    def get_house_by_id(self, id):
        try:
            house = House.objects.get(id=id)
            serializer = self.serializer_class(house)
            return Response(serializer.data)
        except House.DoesNotExist:
            return Response({'detail': 'Дом не найден.'}, status=status.HTTP_404_NOT_FOUND)

    def filter_houses(self, filters, category_name=None, sort_by='priceAsc', title=None):
        filters = filters.copy()
        filters = dict(filters)
        filters = {k: v[0] if isinstance(v, list) and len(v) == 1 else v for k, v in filters.items()}

        houses = House.objects.all()

        if 'category' in filters and filters['category']:
            category_name = filters.pop('category').replace('+', ' ')
            if category_name != 'all':
                try:
                    category = HouseCategory.objects.get(name__iexact=category_name)
                    houses = houses.filter(category=category)
                except HouseCategory.DoesNotExist:
                    return House.objects.none()

        if title:
            houses = houses.filter(title__icontains=title)

        filtered_houses = self.create_dynamic_filter(filters, houses)

        if sort_by == 'priceAsc':
            filtered_houses = filtered_houses.order_by('price')
        elif sort_by == 'priceDesc':
            filtered_houses = filtered_houses.order_by('-price')

        return filtered_houses

    def create_dynamic_filter(self, filters, queryset):
        filters_db = FilterOption.objects.all()
        filter_dict = {}

        cleaned_filters = {
            k: v for k, v in filters.items()
            if v not in ['all', '', None]
        }

        for filter_option in filters_db:
            if filter_option.filter_type == 'exact':
                filter_dict[filter_option.field_name] = django_filters.CharFilter(field_name=filter_option.field_name)
            elif filter_option.filter_type == 'range':
                filter_dict[filter_option.field_name + '__gte'] = django_filters.NumberFilter(
                    field_name=filter_option.field_name, lookup_expr='gte')
                filter_dict[filter_option.field_name + '__lte'] = django_filters.NumberFilter(
                    field_name=filter_option.field_name, lookup_expr='lte')
            elif filter_option.filter_type == 'contains':
                filter_dict[filter_option.field_name] = django_filters.CharFilter(
                    field_name=filter_option.field_name, lookup_expr='icontains')

        house_filter = DynamicHouseFilter(cleaned_filters, queryset=queryset)
        house_filter.filters.update(filter_dict)

        return house_filter.qs





class HouseDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = House.objects.prefetch_related(
        'category',
        'construction_technology',
        'images',
        'interior_images',
        'facade_images',
        'layout_images',
        'finishing_options',
        'documents'
    ).all()
    serializer_class = HouseSerializer

    def get(self, request, *args, **kwargs):
        house = self.get_object()

        cache_key = f"house_detail_{house.id}"
        cached_data = cache.get(cache_key)

        if cached_data:
            return Response(cached_data)

        house_data = HouseSerializer(house).data

        cache.set(cache_key, house_data, timeout=60 * 10)

        return Response(house_data)

    def perform_update(self, serializer):
        super().perform_update(serializer)
        self.clear_cache()

    def perform_destroy(self, instance):
        super().perform_destroy(instance)
        self.clear_cache()

    def clear_cache(self):
        cache_key = f"house_detail_{self.get_object().id}"
        cache.delete(cache_key)


class FilteredHouseListView(APIView):
    serializer_class = HouseSerializer

    def get(self, request):
        filters = request.query_params

        houses = filter_houses(filters)
        serializer = self.serializer_class(houses, many=True)
        return Response(serializer.data)



class ConstructionTechnologyListView(generics.ListCreateAPIView):
    queryset = ConstructionTechnology.objects.all()
    serializer_class = ConstructionTechnologySerializer

    def get_queryset(self):
        options = cache.get('construction_technologies')
        if not options:
            options = super().get_queryset()
            cache.set('construction_technologies', options, timeout=60 * 60)
        return options


class ConstructionTechnologyDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ConstructionTechnology.objects.all()
    serializer_class = ConstructionTechnologySerializer




class HouseCategoryListView(generics.ListCreateAPIView):
    queryset = HouseCategory.objects.prefetch_related('houses')
    serializer_class = HouseCategorySerializer

    def get(self, request, *args, **kwargs):
        cache_key = "house_category_list"
        cached_data = cache.get(cache_key)

        if cached_data:
            return Response(cached_data)


        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        data = serializer.data

        cache.set(cache_key, data, timeout=60 * 60)
        return Response(data)

    def perform_create(self, serializer):
        response = super().perform_create(serializer)
        cache.delete("house_category_list")
        return response


class HouseCategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = HouseCategory.objects.all()
    serializer_class = HouseCategorySerializer
    lookup_field = 'slug'

    def get(self, request, *args, **kwargs):
        category_slug = self.kwargs['slug']
        cache_key = f"house_category_{category_slug}"
        cached_data = cache.get(cache_key)

        if cached_data:
            return Response(cached_data)

        category = get_object_or_404(HouseCategory, slug=category_slug)
        filters = request.query_params
        houses = filter_houses(filters, category=category)

        category_serializer = self.get_serializer(category)
        houses_serializer = HouseSerializer(houses, many=True)

        response_data = {
            "category": category_serializer.data,
            "houses": houses_serializer.data,
        }

        cache.set(cache_key, response_data, timeout=60 * 60)
        return Response(response_data)

    def perform_update(self, serializer):
        response = super().perform_update(serializer)
        cache.delete(f"house_category_{self.kwargs['slug']}")
        cache.delete("house_category_list")
        return response

    def perform_destroy(self, instance):
        cache.delete(f"house_category_{instance.slug}")
        cache.delete("house_category_list")
        super().perform_destroy(instance)


class HouseCategoryDetailByIdView(generics.RetrieveUpdateDestroyAPIView):
    queryset = HouseCategory.objects.all()
    serializer_class = HouseCategorySerializer
    permission_classes = [IsAuthenticated]


class FinishingOptionListView(generics.ListCreateAPIView):
    queryset = FinishingOption.objects.all()
    serializer_class = FinishingOptionSerializer
    #
    # def get(self, request, *args, **kwargs):
    #     cache_key = "finishing_options_list"
    #     cached_data = cache.get(cache_key)
    #
    #     if cached_data:
    #         return Response(cached_data)
    #
    #     queryset = self.get_queryset()
    #     serializer = self.get_serializer(queryset, many=True)
    #     data = serializer.data
    #
    #     cache.set(cache_key, data, timeout=60 * 60)
    #     return Response(data)
    #
    # def perform_create(self, serializer):
    #     response = super().perform_create(serializer)
    #     cache.delete("finishing_options_list")  # Очищаем кэш списка
    #     return response

class FinishingOptionDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = FinishingOption.objects.all()
    serializer_class = FinishingOptionSerializer
    permission_classes = [IsAuthenticated]

    # def get(self, request, *args, **kwargs):
    #     finishing_option = self.get_object()
    #     cache_key = f"finishing_option_{finishing_option.id}"
    #     cached_data = cache.get(cache_key)
    #
    #     if cached_data:
    #         return Response(cached_data)
    #
    #     serializer = self.get_serializer(finishing_option)
    #     data = serializer.data
    #
    #     cache.set(cache_key, data, timeout=60 * 60)
    #     return Response(data)
    #
    # def perform_update(self, serializer):
    #     response = super().perform_update(serializer)
    #     cache.delete(f"finishing_option_{self.get_object().id}")
    #     cache.delete("finishing_options_list")
    #     return response
    #
    # def perform_destroy(self, instance):
    #     cache.delete(f"finishing_option_{instance.id}")
    #     cache.delete("finishing_options_list")
    #     super().perform_destroy(instance)

class DocumentListView(generics.ListCreateAPIView):
    queryset = Document.objects.all()
    serializer_class = DocumentSerializer


class DocumentDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Document.objects.all()
    serializer_class = DocumentSerializer



class ReviewsPagination(PageNumberPagination):
    page_size = 6
    page_size_query_param = 'page_size'
    max_page_size = 100

class ReviewsListView(generics.ListCreateAPIView):
    queryset = Review.objects.all().order_by('-date')
    serializer_class = ReviewSerializer
    parser_classes = [MultiPartParser, FormParser]
    pagination_class = ReviewsPagination

    def get_queryset(self):
        queryset = super().get_queryset()
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        limit = request.query_params.get('limit')
        if limit:
            try:
                limit = int(limit)
                if limit > 0:
                    queryset = queryset[:limit]
                    serializer = self.get_serializer(queryset, many=True)
                    return Response(serializer.data)
            except (ValueError, TypeError):
                pass

        paginator = self.pagination_class()
        result_page = paginator.paginate_queryset(queryset, request)

        serializer = self.get_serializer(result_page, many=True)
        return paginator.get_paginated_response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ReviewsDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Review.objects.all()
    serializer_class = ReviewSerializer
    parser_classes = [MultiPartParser, FormParser]

    def get(self, request, *args, **kwargs):
        review = self.get_object()
        serializer = self.get_serializer(review)
        return Response(serializer.data)

    def update(self, request, *args, **kwargs):
        review = self.get_object()
        files = request.FILES.getlist('uploaded_files')  # Получаем новые файлы
        serializer = self.get_serializer(review, data=request.data, partial=True)

        if serializer.is_valid():
            review = serializer.save()

            for file in files:
                ReviewFile.objects.create(review=review, file=file)

            return Response(serializer.data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()


        ReviewFile.objects.filter(review=instance).delete()

        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)



class OrderListView(generics.ListCreateAPIView):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class OrderDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer


class OrdersByEmailView(APIView):
    def get(self, request):
        email = request.query_params.get('email')
        if not email:
            return Response(
                {"error": "Требуется параметр электронной почты."},
                status=status.HTTP_400_BAD_REQUEST
            )

        orders = Order.objects.filter(email=email)
        if not orders.exists():
            return Response(
                {"error": "На этот адрес электронной почты заказов не найдено."},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = OrderSerializer(orders, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class UserQuestionHouseListView(generics.ListCreateAPIView):
    queryset = UserQuestionHouse.objects.all()
    serializer_class = UserQuestionHouseSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class UserQuestionHouseDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = UserQuestionHouse.objects.all()
    serializer_class = UserQuestionHouseSerializer
    # permission_classes = [IsAuthenticated]


class UserQuestionListView(ListCreateAPIView):
    queryset = UserQuestion.objects.all()
    serializer_class = UserQuestionSerializer

    @method_decorator(cache_page(60 * 15))
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)


class UserQuestionDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = UserQuestion.objects.all()
    serializer_class = UserQuestionSerializer
    # permission_classes = [IsAuthenticated]


class PurchaseHouseListView(generics.ListCreateAPIView):
    queryset = PurchasedHouse.objects.all()
    serializer_class = PurchasedHouseSerializer

    def get(self, request, *args, **kwargs):
        cache_key = "purchase_house_list"
        cached_data = cache.get(cache_key)

        if cached_data:
            return Response(cached_data)

        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        data = serializer.data

        cache.set(cache_key, data, timeout=60 * 60)
        return Response(data)

    def perform_create(self, serializer):
        response = super().perform_create(serializer)
        cache.delete("purchase_house_list")
        return response

    def get_queryset(self):
        construction_status = self.request.query_params.get('construction_status', None)

        queryset = PurchasedHouse.objects.all().select_related('house')
        queryset = queryset.prefetch_related(
            'house__images', 'house__interior_images', 'house__facade_images',
            'house__layout_images', 'house__category', 'house__construction_technology',
            'house__documents', 'house__finishing_options'
        )

        if construction_status:
            return queryset.filter(construction_status=construction_status)
        return queryset


class PurchaseHouseDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = PurchasedHouse.objects.all()
    serializer_class = PurchasedHouseSerializer
    # permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        house_id = self.kwargs['pk']
        cache_key = f"purchase_house_{house_id}"
        cached_data = cache.get(cache_key)

        if cached_data:
            return Response(cached_data)

        house = self.get_object()
        serializer = self.get_serializer(house)
        data = serializer.data

        cache.set(cache_key, data, timeout=60 * 60)
        return Response(data)

    def perform_update(self, serializer):
        response = super().perform_update(serializer)
        cache.delete(f"purchase_house_{self.kwargs['pk']}")
        cache.delete("purchase_house_list")
        return response

    def perform_destroy(self, instance):
        cache.delete(f"purchase_house_{instance.pk}")
        cache.delete("purchase_house_list")
        super().perform_destroy(instance)



class FilterOptionListView(generics.ListCreateAPIView):
    queryset = FilterOption.objects.all()
    serializer_class = FilterOptionsSerializer


class FilterOptionDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = FilterOption.objects.all()
    serializer_class = FilterOptionsSerializer
    permission_classes = [IsAuthenticated]


class CreateHouseAPIView(APIView):
    # permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, format=None):
        serializer = HouseSerializer(data=request.data)
        if serializer.is_valid():
            house = serializer.save()

            for file in request.FILES.getlist('images'):
                img = Image(image=file)
                img.save()
                house.images.add(img)

            for file in request.FILES.getlist('interior_images'):
                img = Image(image=file)
                img.save()
                house.interior_images.add(img)

            for file in request.FILES.getlist('facade_images'):
                img = Image(image=file)
                img.save()
                house.facade_images.add(img)

            for file in request.FILES.getlist('layout_images'):
                img = Image(image=file)
                img.save()
                house.layout_images.add(img)

            for file in request.FILES.getlist('documents'):
                document = Document(file=file)
                document.save()
                house.documents.add(document)

            return Response({"message": "Дом успешно создан"}, status=201)
        return Response(serializer.errors, status=400)


class UpdateHouseAPIView(APIView):
    # permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def patch(self, request, house_id, format=None):
        try:
            house = House.objects.get(id=house_id)
        except House.DoesNotExist:
            return Response({"message": "Дом не найден"}, status=404)

        serializer = HouseSerializer(house, data=request.data, partial=True)
        if serializer.is_valid():
            house = serializer.save()

            if 'remove_images' in request.data:
                remove_images = request.data.getlist('remove_images')
                for image_id in remove_images:
                    try:
                        image = Image.objects.get(id=image_id)
                        house.images.remove(image)
                        image.delete()
                    except Image.DoesNotExist:
                        continue

            for file in request.FILES.getlist('images'):
                img = Image(image=file)
                img.save()
                house.images.add(img)

            for file in request.FILES.getlist('interior_images'):
                img = Image(image=file)
                img.save()
                house.interior_images.add(img)

            for file in request.FILES.getlist('facade_images'):
                img = Image(image=file)
                img.save()
                house.facade_images.add(img)

            for file in request.FILES.getlist('layout_images'):
                img = Image(image=file)
                img.save()
                house.layout_images.add(img)

            if 'remove_documents' in request.data:
                remove_documents = request.data.getlist('remove_documents')
                for document_id in remove_documents:
                    try:
                        document = Document.objects.get(id=document_id)
                        house.documents.remove(document)
                        document.delete()
                    except Document.DoesNotExist:
                        continue

            for file in request.FILES.getlist('documents'):
                document = Document(file=file)
                document.save()
                house.documents.add(document)

            return Response({"message": "Дом успешно обновлен"}, status=200)

        return Response(serializer.errors, status=400)


class DeleteImageView(APIView):
    # permission_classes = [IsAuthenticated]
    def delete(self, request, house_id, image_id, category):
        house = get_object_or_404(House, id=house_id)
        image = get_object_or_404(Image, id=image_id)


        if category == 'images':
            if image in house.images.all():
                house.images.remove(image)
        elif category == 'interior_images':
            if image in house.interior_images.all():
                house.interior_images.remove(image)
        elif category == 'facade_images':
            if image in house.facade_images.all():
                house.facade_images.remove(image)
        elif category == 'layout_images':
            if image in house.layout_images.all():
                house.layout_images.remove(image)
        else:
            return JsonResponse({'status': 'error', 'message': 'Неправильная категория изображения'}, status=400)


        if image.houses.count() == 0:
            image.delete()

        return JsonResponse({'status': 'success', 'message': 'Картинка успешно удалена'})

class DeleteDocumentView(APIView):
    # permission_classes = [IsAuthenticated]

    def delete(self, request, house_id, document_id):
        house = get_object_or_404(House, id=house_id)
        document = get_object_or_404(Document, id=document_id)

        if document in house.documents.all():
            house.documents.remove(document)
        else:
            return Response({'message': 'Документ не связан с этим домом'}, status=400)


        if document.houses.count() == 0:
            document.delete()

        return Response({'message': 'Документ успешно удален'}, status=200)


def export_orders_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = [
        "ID", "Название дома", "Покупатель", "Телефон", "Email",
        "Место строительства", "Отделка", "Дата заказа", "Статус"
    ]
    ws.append(headers)

    orders = Order.objects.all()

    for order in orders:
        row = [
            order.id,
            order.house.title if order.house else "Не указано",
            order.name,
            order.phone,
            order.email,
            order.construction_place,
            order.finishing_option.title if order.finishing_option else "Не указано",
            order.data_created.strftime('%Y-%m-%d'),
            order.status
        ]
        ws.append(row)


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'

    wb.save(response)

    return response

def export_purchased_houses(request):
    purchased_houses = PurchasedHouse.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchased Houses"

    ws.append([
        "Дом",
        "Дата покупки",
        "Покупатель",
        "Телефон",
        "Почта",
        "Статус строительства",
        "Широта",
        "Долгота",
    ])

    for house in purchased_houses:
        ws.append([
            house.house.title,
            house.purchase_date,
            house.buyer_name,
            house.buyer_phone,
            house.buyer_email,
            house.construction_status,
            house.latitude if house.latitude else "Не указано",
            house.longitude if house.longitude else "Не указано",
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="purchased_houses.xlsx"'

    wb.save(response)
    return response

def export_user_questions_and_houses(request):
    user_questions = UserQuestion.objects.all()
    user_question_houses = UserQuestionHouse.objects.all()


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Questions and Houses"

    ws.append([
        "Имя",
        "Телефон",
        "Дата создания",
        "Статус"
    ])

    for question in user_questions:
        ws.append([
            question.name,
            question.phone,
            question.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            question.status,
        ])


    ws.append([])

    ws.append([
        "Имя",
        "Телефон",
        "Email",
        "Дом",
        "Вопрос",
        "Дата создания",
        "Статус"
    ])

    for question_house in user_question_houses:
        ws.append([
            question_house.name,
            question_house.phone,
            question_house.email,
            question_house.house.title,
            question_house.question,
            question_house.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            question_house.status,
        ])


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="user_questions_and_houses.xlsx"'

    wb.save(response)
    return response

class BlogListCreateView(generics.ListCreateAPIView):
    queryset = Blog.objects.select_related('category').all().order_by('-date')
    serializer_class = BlogSerializer
    parser_classes = [MultiPartParser, FormParser]
    pagination_class = Pagination

    def get_queryset(self):
        queryset = super().get_queryset()

        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category__name__iexact=category)


        return queryset


class BlogDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Blog.objects.all()
    serializer_class = BlogSerializer
    parser_classes = [MultiPartParser, FormParser]

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

class BlogCategoryListView(generics.ListCreateAPIView):
    queryset = BlogCategory.objects.all()
    serializer_class = BlogCategorySerializer

class BlogsByCategoryView(generics.ListAPIView):
    serializer_class = BlogSerializer

    def get_queryset(self):
        category_id = self.kwargs['category_id']
        return Blog.objects.filter(category_id=category_id)